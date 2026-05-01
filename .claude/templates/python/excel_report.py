"""
Template: Geração de Relatório Excel Corporativo
Biblioteca: openpyxl
Adaptar: dados, abas, colunas, KPIs e formatação
Padrão: Config → Raw → Calc → Dashboard
"""

import os
from datetime import date, datetime
from typing import Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint


# ---------------------------------------------------------------------------
# Paleta corporativa — ajustar conforme identidade visual
# ---------------------------------------------------------------------------
COLORS = {
    "header_bg":   "1A3A5C",   # azul escuro
    "header_font": "FFFFFF",   # branco
    "row_even":    "EBF1F5",   # azul clarinho
    "row_odd":     "FFFFFF",   # branco
    "positive":    "22C55E",   # verde
    "negative":    "EF4444",   # vermelho
    "warning":     "F59E0B",   # amarelo
    "accent":      "3B82F6",   # azul
    "kpi_bg":      "F0F9FF",   # azul muito claro para cards
}

FONT_TITLE   = Font(name="Calibri", size=14, bold=True, color=COLORS["header_font"])
FONT_HEADER  = Font(name="Calibri", size=10, bold=True, color=COLORS["header_font"])
FONT_BODY    = Font(name="Calibri", size=10)
FONT_KPI     = Font(name="Calibri", size=20, bold=True, color=COLORS["accent"])

FILL_HEADER  = PatternFill("solid", fgColor=COLORS["header_bg"])
FILL_EVEN    = PatternFill("solid", fgColor=COLORS["row_even"])
FILL_KPI     = PatternFill("solid", fgColor=COLORS["kpi_bg"])

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Helpers de formatação
# ---------------------------------------------------------------------------
def style_header_row(ws, row: int, col_start: int, col_end: int) -> None:
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font   = FONT_HEADER
        cell.fill   = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def style_data_rows(ws, row_start: int, row_end: int,
                    col_start: int, col_end: int) -> None:
    for row in range(row_start, row_end + 1):
        fill = FILL_EVEN if row % 2 == 0 else PatternFill("solid", fgColor=COLORS["row_odd"])
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.font      = FONT_BODY
            cell.fill      = fill
            cell.alignment = ALIGN_LEFT
            cell.border    = BORDER_THIN


def autofit_columns(ws, min_width: int = 10, max_width: int = 40) -> None:
    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            max(min_width, min(length + 2, max_width))


def freeze_header(ws, row: int = 2, col: int = 1) -> None:
    ws.freeze_panes = ws.cell(row=row, column=col)


# ---------------------------------------------------------------------------
# Aba: Config — parâmetros e metadados do relatório
# ---------------------------------------------------------------------------
def create_config_sheet(wb: Workbook, params: dict) -> None:
    ws = wb.create_sheet("Config")
    ws.sheet_state = "hidden"  # oculta para o usuário final

    ws["A1"] = "Parâmetro"
    ws["B1"] = "Valor"
    style_header_row(ws, 1, 1, 2)

    for i, (key, value) in enumerate(params.items(), start=2):
        ws.cell(row=i, column=1, value=key).font  = FONT_BODY
        ws.cell(row=i, column=2, value=value).font = FONT_BODY

    autofit_columns(ws)


# ---------------------------------------------------------------------------
# Aba: Raw — dados brutos importados (não editar manualmente)
# ---------------------------------------------------------------------------
def create_raw_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str = "Raw_Dados") -> None:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_state = "veryHidden"  # protege dados brutos

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = FONT_HEADER
        cell.fill   = FILL_HEADER
        cell.alignment = ALIGN_CENTER

    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value).font = FONT_BODY

    autofit_columns(ws)


# ---------------------------------------------------------------------------
# Aba: Dashboard — KPI cards + tabela + gráfico
# ---------------------------------------------------------------------------
def create_dashboard_sheet(wb: Workbook, kpis: list[dict],
                            df_summary: pd.DataFrame, title: str) -> None:
    ws = wb.create_sheet("Dashboard")

    # ── Título ────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value     = title
    title_cell.font      = Font(name="Calibri", size=16, bold=True, color=COLORS["header_bg"])
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 30

    subtitle_cell = ws["A2"]
    ws.merge_cells("A2:H2")
    subtitle_cell.value     = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    subtitle_cell.font      = Font(name="Calibri", size=9, color="888888")
    subtitle_cell.alignment = ALIGN_CENTER

    # ── KPI Cards (linha 4) ────────────────────────────────────
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 35
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 18

    for i, kpi in enumerate(kpis[:4]):  # máximo 4 KPI cards
        col = i * 2 + 1  # colunas A, C, E, G
        col_letter_start = get_column_letter(col)
        col_letter_end   = get_column_letter(col + 1)

        ws.merge_cells(f"{col_letter_start}4:{col_letter_end}4")
        ws.merge_cells(f"{col_letter_start}5:{col_letter_end}5")
        ws.merge_cells(f"{col_letter_start}6:{col_letter_end}6")
        ws.merge_cells(f"{col_letter_start}7:{col_letter_end}7")

        # Label
        label_cell = ws[f"{col_letter_start}4"]
        label_cell.value     = kpi.get("label", "")
        label_cell.font      = Font(name="Calibri", size=9, bold=True, color="888888")
        label_cell.alignment = ALIGN_CENTER
        label_cell.fill      = FILL_KPI

        # Valor
        value_cell = ws[f"{col_letter_start}5"]
        value_cell.value     = kpi.get("value", "")
        value_cell.font      = FONT_KPI
        value_cell.alignment = ALIGN_CENTER
        value_cell.fill      = FILL_KPI

        # Delta
        delta = kpi.get("delta", 0)
        delta_cell = ws[f"{col_letter_start}6"]
        delta_cell.value     = f"{'▲' if delta >= 0 else '▼'} {abs(delta):.1f}%"
        delta_cell.font      = Font(
            name="Calibri", size=10, bold=True,
            color=COLORS["positive"] if delta >= 0 else COLORS["negative"]
        )
        delta_cell.alignment = ALIGN_CENTER
        delta_cell.fill      = FILL_KPI

        # Contexto
        ctx_cell = ws[f"{col_letter_start}7"]
        ctx_cell.value     = kpi.get("context", "")
        ctx_cell.font      = Font(name="Calibri", size=8, color="888888")
        ctx_cell.alignment = ALIGN_CENTER
        ctx_cell.fill      = FILL_KPI

    # ── Tabela resumo (linha 10+) ──────────────────────────────
    table_start_row = 10
    ws[f"A{table_start_row - 1}"] = "Resumo por Dimensão"
    ws[f"A{table_start_row - 1}"].font = Font(name="Calibri", size=11, bold=True,
                                               color=COLORS["header_bg"])

    headers = list(df_summary.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=table_start_row, column=col_idx, value=header)
    style_header_row(ws, table_start_row, 1, len(headers))

    for row_idx, row in enumerate(
        dataframe_to_rows(df_summary, index=False, header=False),
        start=table_start_row + 1
    ):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    style_data_rows(ws, table_start_row + 1,
                    table_start_row + len(df_summary),
                    1, len(headers))
    autofit_columns(ws)
    freeze_header(ws, row=table_start_row + 1)


# ---------------------------------------------------------------------------
# Gráfico de barras — adicionar em qualquer aba
# ---------------------------------------------------------------------------
def add_bar_chart(ws, title: str, data_ref: Reference,
                  cats_ref: Reference, anchor: str = "J4") -> None:
    chart = BarChart()
    chart.type   = "col"
    chart.title  = title
    chart.y_axis.title = "Valor"
    chart.x_axis.title = ""
    chart.style  = 10
    chart.width  = 18
    chart.height = 12

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, anchor)


# ---------------------------------------------------------------------------
# Orquestrador principal — adaptar conforme o relatório
# ---------------------------------------------------------------------------
def build_report(
    df_raw: pd.DataFrame,
    df_summary: pd.DataFrame,
    kpis: list[dict],
    output_path: str,
    title: str = "Relatório Executivo",
    reference_date: date | None = None,
) -> str:
    reference_date = reference_date or date.today()

    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão

    # Parâmetros do relatório
    params = {
        "titulo":            title,
        "data_referencia":   reference_date.strftime("%d/%m/%Y"),
        "gerado_em":         datetime.now().strftime("%d/%m/%Y %H:%M"),
        "total_registros":   len(df_raw),
    }

    create_config_sheet(wb, params)
    create_raw_sheet(wb, df_raw, "Raw_Dados")
    create_dashboard_sheet(wb, kpis, df_summary, title)

    wb.save(output_path)
    print(f"Relatório gerado: {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# Exemplo de uso — substituir pelos dados reais
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import numpy as np

    # Dados simulados — substituir por query real
    df_raw = pd.DataFrame({
        "regional":    np.random.choice(["Sul", "Sudeste", "Norte"], 100),
        "canal":       np.random.choice(["Direto", "Distribuidor"], 100),
        "receita":     np.random.uniform(1000, 50000, 100).round(2),
        "meta":        np.random.uniform(1000, 50000, 100).round(2),
        "dt_venda":    pd.date_range("2026-01-01", periods=100, freq="D")[:100],
    })

    df_summary = df_raw.groupby("regional").agg(
        receita=("receita", "sum"),
        meta=("meta", "sum"),
        pedidos=("receita", "count"),
    ).reset_index()
    df_summary["atingimento_pct"] = (df_summary["receita"] / df_summary["meta"] * 100).round(1)

    kpis = [
        {"label": "Receita Total",  "value": "R$ 2,8M",  "delta": +8.3,  "context": "vs meta: +5%"},
        {"label": "Margem Bruta",   "value": "38,4%",    "delta": -1.2,  "context": "meta: 40%"},
        {"label": "Pedidos",        "value": "1.842",    "delta": +12.1, "context": "ticket: R$ 1.520"},
        {"label": "Atingimento",    "value": "105%",     "delta": +5.0,  "context": "acima da meta"},
    ]

    build_report(
        df_raw=df_raw,
        df_summary=df_summary,
        kpis=kpis,
        output_path="relatorio_executivo.xlsx",
        title="Performance Comercial — Jan/2026",
    )
